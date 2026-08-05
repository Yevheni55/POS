"""Obnovi sezonny P&L xlsx na ploche. Spusta sa Windows Task Schedulerom o 04:00.

Cely tok: ssh na kasu -> node vnutri kontajnera pos-app-1 -> JSON -> xlsx.

Bezpecnostne zasady tohto skriptu:
  * Ked cokolvek zlyha (kasa nedostupna, VPN dole, Docker stoji), povodny xlsx
    sa NEPREPISE. Stary spravny subor je vzdy lepsi nez novy prazdny.
  * Pred kazdym zapisom sa spravi zaloha do podadresara `archiv/`, nie na plochu.
  * Vsetko sa loguje do `logs/refresh.log` — po nocnom behu sa da spatne zistit,
    co sa stalo.

Usage:  python refresh_season_xlsx.py
Env:    SEASON_XLSX_OUT  cielovy subor (default: plocha)
        DEPLOY_HOST      ssh host kasy (default: surfs@100.95.64.38)
"""
from __future__ import annotations

import datetime as dt
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent
OUT = Path(os.environ.get("SEASON_XLSX_OUT")
           or (Path.home() / "Desktop" / "SurfSpirit-Sezona-2026.xlsx"))
HOST = os.environ.get("DEPLOY_HOST", "surfs@100.95.64.38")
ARCHIV = HERE / "archiv"
LOGDIR = HERE / "logs"
KEEP_BACKUPS = 14
SSH_OPTS = ["-o", "ConnectTimeout=25", "-o", "BatchMode=yes"]

SK_DAY = ["Po", "Ut", "St", "Št", "Pi", "So", "Ne"]
SK_MONTH = ["", "Január", "Február", "Marec", "Apríl", "Máj", "Jún",
            "Júl", "August", "September", "Október", "November", "December"]


def log(msg: str) -> None:
    LOGDIR.mkdir(parents=True, exist_ok=True)
    line = f"{dt.datetime.now():%Y-%m-%d %H:%M:%S}  {msg}"
    print(line)
    with (LOGDIR / "refresh.log").open("a", encoding="utf-8") as fh:
        fh.write(line + "\n")


def run(cmd: list[str], timeout: int = 300) -> str:
    p = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8",
                       errors="replace", timeout=timeout)
    if p.returncode != 0:
        raise RuntimeError(f"{cmd[0]} skoncil s {p.returncode}: {(p.stderr or '').strip()[:400]}")
    return p.stdout or ""


def between(text: str, start: str, end: str):
    i, j = text.find(start), text.find(end)
    if i == -1 or j == -1:
        raise RuntimeError(f"vo vystupe chyba blok {start}")
    return json.loads(text[i + len(start):j].strip())


def fetch() -> tuple[dict, dict]:
    """Nahra dump skript na kasu, pusti ho v kontajneri, vrati (payload, vat)."""
    run(["scp", *SSH_OPTS, str(HERE / "dump-season.mjs"), f"{HOST}:season-dump.mjs"], 120)
    out = run(["ssh", *SSH_OPTS, HOST,
               "docker cp season-dump.mjs pos-app-1:/app/server/season-dump.mjs && "
               "docker exec -w /app/server pos-app-1 node season-dump.mjs"], 300)
    payload = between(out, "PAYLOAD_START", "PAYLOAD_END")
    payload["firmy"] = between(out, "FIRMY_START", "FIRMY_END")
    return payload, between(out, "VAT_START", "VAT_END")


def build(payload: dict, vat: dict, target: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    FONT = "Calibri"
    EUR = '#,##0.00" €"'
    BOLD = Font(name=FONT, size=11, bold=True)
    PLAIN = Font(name=FONT, size=11)
    SMALL = Font(name=FONT, size=9, italic=True, color="595959")

    daily = payload["daily"]
    if not daily:
        raise RuntimeError("summary nevratil ziadne denne riadky — nezapisujem")
    rev = float(payload["totalRevenue"])
    cogs = float(payload["totalCogs"])
    labor = float(payload["totalLabor"])
    meal = float(payload["totalStaffMeal"])
    odpis = float(payload["totalOdpis"])
    shisha_rev = float(payload["shisha"]["revenue"])
    shisha_cnt = payload["shisha"].get("count", "")
    vat_real = float(vat["vatTotal"])
    profit = round(rev - cogs - labor - meal, 2)

    first = dt.date.fromisoformat(daily[0]["date"])
    last = dt.date.fromisoformat(daily[-1]["date"])
    sk = lambda d: f"{d.day}.{d.month}.{d.year}"  # noqa: E731
    pct = lambda v: f"{v / rev * 100:.1f} %" if rev else "—"  # noqa: E731

    wb = Workbook()
    ws = wb.active
    ws.title = "Sezóna 2026"

    def put(row, values, bold=False, fmts=None):
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=i + 1, value=v)
            c.font = BOLD if bold else PLAIN
            if fmts and i < len(fmts) and fmts[i]:
                c.number_format = fmts[i]

    ws.cell(row=1, column=1,
            value=f"SurfSpirit — Sezóna {first.year} ({sk(first)} – {sk(last)})"
            ).font = Font(name=FONT, size=13, bold=True)
    ws.cell(row=2, column=1, value=(
        f"Automaticky aktualizované {dt.datetime.now():%d.%m.%Y %H:%M}"
        " · zdroj: report Sezóna (tržby = fiškálne platby + shisha + odpis;"
        " výsledok = tržby − výroba − mzdy − zam. spotreba). Čísla sú BRUTTO."
    )).font = SMALL

    put(4, ["SÚHRN ZA SEZÓNU", "€", "% z tržieb"], bold=True)
    put(5, ["Celkové tržby", rev, "", f"z toho odpis {odpis} €, shisha {shisha_rev} €"],
        bold=True, fmts=[None, EUR])
    put(6, ["Náklady na výrobu", cogs, pct(cogs), "suroviny podľa receptúr"], fmts=[None, EUR])
    put(7, ["Mzdy", labor, pct(labor), "dochádzka × hodinovka"], fmts=[None, EUR])
    put(8, ["Zamestnanecká spotreba", meal, pct(meal), "suroviny staff meals"], fmts=[None, EUR])
    put(9, ["VÝSLEDOK", profit, f"{profit / rev * 100:.1f} % marža"], bold=True, fmts=[None, EUR])
    put(10, ["DPH skutočne odvedená", round(vat_real, 2), pct(vat_real),
             "zo skutočných fiškálnych dokladov; obdobie neplatiteľa nesie 0 %"],
        fmts=[None, EUR])

    # --- rozpad podla danoveho subjektu -----------------------------------
    # Iba fiskalne trzby: shisha a odpis doklad nemaju, takze sa k firme priradit
    # nedaju a tento blok sa NEROVNA riadku "Celkove trzby". DPH plati len
    # SL management — predchadzajuce dva subjekty boli neplatitelia a ich doklady
    # realne odisli s 0 %.
    NAMES = {"54588481": "SL management, s.r.o.",
             "57513708": "Švískej s. r. o.",
             "57307512": "Prvý subjekt"}
    firmy = payload.get("firmy") or []
    put(12, ["PO FIRMÁCH (tržby cez eKasu)", "IČO", "Obdobie", "Účty", "Tržby",
             "DPH odvedená"], bold=True)
    fr = 13
    for f in firmy:
        ico = str(f.get("ico") or "")
        d_od = dt.date.fromisoformat(f["od"])
        d_do = dt.date.fromisoformat(f["do"])
        obdobie = f"{d_od.day}.{d_od.month}. – {d_do.day}.{d_do.month}."
        dph = round(float(f["dph"]), 2)
        put(fr, [NAMES.get(ico, f"IČO {ico}"), ico, obdobie, f["uctov"],
                 round(float(f["trzba"]), 2), dph if dph else "0,00 (neplatiteľ)"],
            fmts=[None, None, None, "0", EUR, EUR if dph else None])
        fr += 1
    put(fr, ["SPOLU", "", "", f"=SUM(D13:D{fr - 1})", f"=SUM(E13:E{fr - 1})",
             f"=SUM(F13:F{fr - 1})"], bold=True, fmts=[None, None, None, "0", EUR, EUR])
    fr += 1
    ws.cell(row=fr, column=1, value=(
        "Bez shishy a odpisu — tie fiškálny doklad nemajú, k firme sa priradiť nedajú. "
        "Mzdy sa podľa firmy nedelia vôbec (sú to náklady prevádzky).")).font = SMALL
    fr += 2

    put(fr, ["PO MESIACOCH", "Tržby", "Výroba", "Mzdy", "Zam. spotreba", "Výsledok",
             "Aktívnych dní"], bold=True)
    month_header = fr
    months: dict[dt.date, dict] = {}
    for d in daily:
        key = dt.date.fromisoformat(d["date"]).replace(day=1)
        m = months.setdefault(key, {"rev": 0.0, "cogs": 0.0, "labor": 0.0,
                                    "meal": 0.0, "days": 0})
        m["rev"] += d["revenue"]
        m["cogs"] += d["cogs"]
        m["labor"] += d["labor"]
        m["meal"] += d["staffMeal"]
        m["days"] += 1

    r = month_header + 1
    for key in sorted(months):
        m = months[key]
        label = SK_MONTH[key.month]
        if key == first.replace(day=1):
            label += f" (od {first.day}.{first.month}.)"
        if key == last.replace(day=1):
            label += f" (do {last.day}.{last.month}.)"
        p = m["rev"] - m["cogs"] - m["labor"] - m["meal"]
        put(r, [label, round(m["rev"], 2), round(m["cogs"], 2), round(m["labor"], 2),
                round(m["meal"], 2), round(p, 2), m["days"]],
            fmts=[None, EUR, EUR, EUR, EUR, EUR, "0"])
        r += 1
    put(r, [f"Shisha (celá sezóna, {shisha_cnt} ks — nie je v denných riadkoch)",
            shisha_rev, 0, 0, 0, shisha_rev, ""], fmts=[None, EUR, EUR, EUR, EUR, EUR])
    m_first, m_last = month_header + 1, r
    r += 1
    put(r, ["SPOLU"] + [f"=SUM({c}{m_first}:{c}{m_last})" for c in "BCDEFG"],
        bold=True, fmts=[None, EUR, EUR, EUR, EUR, EUR, "0"])
    r += 2

    put(r, ["PO DŇOCH"], bold=True)
    r += 1
    put(r, ["Dátum", "Deň", "Účty", "Tržby", "Výroba", "Mzdy", "Zam. spotreba", "Výsledok"],
        bold=True)
    header = r
    r += 1
    d_first = r
    for d in daily:
        dd = dt.date.fromisoformat(d["date"])
        p = d["revenue"] - d["cogs"] - d["labor"] - d["staffMeal"]
        put(r, [sk(dd), SK_DAY[dd.weekday()], d["orders"], round(d["revenue"], 2),
                round(d["cogs"], 2), round(d["labor"], 2), round(d["staffMeal"], 2),
                round(p, 2)],
            fmts=[None, None, "0", EUR, EUR, EUR, EUR, EUR])
        r += 1
    d_last = r - 1
    put(r, ["SPOLU (bez shisha)", ""]
        + [f"=SUM({c}{d_first}:{c}{d_last})" for c in "CDEFGH"],
        bold=True, fmts=[None, None, "0", EUR, EUR, EUR, EUR, EUR])

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 15
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = f"A{header + 1}"

    # Zapis cez docasny subor + atomicky presun: ked by openpyxl padol uprostred,
    # na ploche neostane orezany subor.
    # mkstemp vracia OTVORENY deskriptor — na Windows drzi subor zamknuty a
    # os.replace() by potom padol na PermissionError, co vyzera ako "otvorene
    # v Exceli". Preto ho hned zatvarame.
    fd, tmp_name = tempfile.mkstemp(suffix=".xlsx", dir=str(target.parent))
    os.close(fd)
    tmp = Path(tmp_name)
    try:
        wb.save(tmp)
        os.replace(tmp, target)
    except BaseException:
        tmp.unlink(missing_ok=True)
        raise
    log(f"zapisane: trzby {rev:.2f} | vyroba {cogs:.2f} | mzdy {labor:.2f} | "
        f"vysledok {profit:.2f} | DPH {vat_real:.2f} | dni {len(daily)}")


def backup_and_prune() -> None:
    if not OUT.exists():
        return
    ARCHIV.mkdir(parents=True, exist_ok=True)
    stamp = dt.datetime.fromtimestamp(OUT.stat().st_mtime).strftime("%Y-%m-%d-%H%M")
    bak = ARCHIV / f"{OUT.stem}-{stamp}.xlsx"
    if not bak.exists():
        shutil.copy2(OUT, bak)
    olds = sorted(ARCHIV.glob(f"{OUT.stem}-*.xlsx"))
    for old in olds[:-KEEP_BACKUPS]:
        old.unlink(missing_ok=True)


def main() -> int:
    try:
        payload, vat = fetch()
    except Exception as exc:  # kasa nedostupna / Docker stoji / VPN dole
        log(f"CHYBA pri stahovani dat: {exc}")
        log("povodny subor ostava nezmeneny")
        return 1
    try:
        backup_and_prune()
        build(payload, vat, OUT)
    except PermissionError:
        log(f"CHYBA: {OUT.name} je otvoreny v Exceli — zavri ho, obnova sa preskakuje")
        return 2
    except Exception as exc:
        log(f"CHYBA pri stavbe xlsx: {exc}")
        return 3
    return 0


if __name__ == "__main__":
    sys.exit(main())
